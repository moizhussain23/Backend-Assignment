from celery import shared_task
import openpyxl
from datetime import datetime
from .models import Customer, Loan
from django.db import transaction
import logging

logger = logging.getLogger(__name__)

@shared_task
def ingest_customer_data():
    """
    Background task to ingest customer data from Excel file
    """
    try:
        wb = openpyxl.load_workbook('customer_data.xlsx')
        sheet = wb.active
        
        customers_created = 0
        customers_updated = 0
        
        with transaction.atomic():
            for row_num in range(2, sheet.max_row + 1):  # Skip header row
                row_data = [cell.value for cell in sheet[row_num]]
                
                if len(row_data) >= 7 and row_data[0] is not None:
                    customer_id = row_data[0]
                    first_name = row_data[1]
                    last_name = row_data[2]
                    age = row_data[3]
                    phone_number = row_data[4]
                    monthly_salary = row_data[5]
                    approved_limit = row_data[6]
                    
                    # Try to get existing customer or create new one
                    customer, created = Customer.objects.get_or_create(
                        customer_id=customer_id,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'age': age,
                            'phone_number': phone_number,
                            'monthly_salary': monthly_salary,
                            'approved_limit': approved_limit,
                            'current_debt': 0  # Default value
                        }
                    )
                    
                    if created:
                        customers_created += 1
                    else:
                        # Update existing customer
                        customer.first_name = first_name
                        customer.last_name = last_name
                        customer.age = age
                        customer.phone_number = phone_number
                        customer.monthly_salary = monthly_salary
                        customer.approved_limit = approved_limit
                        customer.save()
                        customers_updated += 1
        
        logger.info(f"Customer data ingestion completed: {customers_created} created, {customers_updated} updated")
        return f"Successfully processed customer data: {customers_created} created, {customers_updated} updated"
        
    except Exception as e:
        logger.error(f"Error ingesting customer data: {str(e)}")
        raise e


@shared_task
def ingest_loan_data():
    """
    Background task to ingest loan data from Excel file
    """
    try:
        wb = openpyxl.load_workbook('loan_data.xlsx')
        sheet = wb.active
        
        loans_created = 0
        loans_updated = 0
        
        with transaction.atomic():
            for row_num in range(2, sheet.max_row + 1):  # Skip header row
                row_data = [cell.value for cell in sheet[row_num]]
                
                if len(row_data) >= 9 and row_data[0] is not None:
                    customer_id = row_data[0]
                    loan_id = row_data[1]
                    loan_amount = row_data[2]
                    tenure = row_data[3]
                    interest_rate = row_data[4]
                    monthly_payment = row_data[5]
                    emis_paid_on_time = row_data[6]
                    start_date = row_data[7]
                    end_date = row_data[8]
                    
                    # Convert datetime objects to date if needed
                    if isinstance(start_date, datetime):
                        start_date = start_date.date()
                    if isinstance(end_date, datetime):
                        end_date = end_date.date()
                    
                    try:
                        customer = Customer.objects.get(customer_id=customer_id)
                        
                        loan, created = Loan.objects.get_or_create(
                            loan_id=loan_id,
                            defaults={
                                'customer': customer,
                                'loan_amount': loan_amount,
                                'tenure': tenure,
                                'interest_rate': interest_rate,
                                'monthly_repayment': monthly_payment,
                                'emis_paid_on_time': emis_paid_on_time,
                                'start_date': start_date,
                                'end_date': end_date,
                            }
                        )
                        
                        if created:
                            loans_created += 1
                        else:
                            # Update existing loan
                            loan.customer = customer
                            loan.loan_amount = loan_amount
                            loan.tenure = tenure
                            loan.interest_rate = interest_rate
                            loan.monthly_repayment = monthly_payment
                            loan.emis_paid_on_time = emis_paid_on_time
                            loan.start_date = start_date
                            loan.end_date = end_date
                            loan.save()
                            loans_updated += 1
                            
                    except Customer.DoesNotExist:
                        logger.warning(f"Customer with ID {customer_id} not found for loan {loan_id}")
                        continue
        
        logger.info(f"Loan data ingestion completed: {loans_created} created, {loans_updated} updated")
        return f"Successfully processed loan data: {loans_created} created, {loans_updated} updated"
        
    except Exception as e:
        logger.error(f"Error ingesting loan data: {str(e)}")
        raise e


@shared_task
def calculate_current_debt():
    """
    Background task to calculate and update current debt for all customers
    """
    try:
        updated_customers = 0
        
        for customer in Customer.objects.all():
            # Calculate current debt based on active loans
            active_loans = customer.loans.filter(end_date__gte=datetime.now().date())
            current_debt = sum(loan.loan_amount - (loan.emis_paid_on_time * loan.monthly_repayment) 
                             for loan in active_loans)
            current_debt = max(0, current_debt)  # Ensure non-negative
            
            customer.current_debt = current_debt
            customer.save()
            updated_customers += 1
        
        logger.info(f"Current debt calculation completed for {updated_customers} customers")
        return f"Successfully updated current debt for {updated_customers} customers"
        
    except Exception as e:
        logger.error(f"Error calculating current debt: {str(e)}")
        raise e